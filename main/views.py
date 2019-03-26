"""
Views here are mostly subclasses of the Django generic CreateView. Each view has
its own template, named in snake-case for the model to-be-created. Each view
presents a Django form based on a model.

The views are linked together in a linear sequence that reflects the
dependencies of the data. Successfully submitting one form redirects the user to
the next form. Each model has a foreign key into the preceding model. One
exception is PrimerDesign which depends on GuideSelection, two steps back in the
sequence.
"""
import copy
import logging
import os
import time
import urllib

from concurrent.futures import ThreadPoolExecutor
from io import BytesIO, StringIO
from itertools import islice
from typing import Any, Optional

import openpyxl  # type: ignore
import sample_sheet as illumina  # type: ignore

from django.http import Http404
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseRedirect
from django.shortcuts import render
from django.views import View
from django.views.generic import DetailView, ListView
from django.views.generic.detail import SingleObjectMixin
from django.views.generic.edit import CreateView
from django.core.exceptions import ObjectDoesNotExist

import webscraperequest

from crispresso.fastqs import find_matching_pairs, reverse_complement
from crispresso.s3 import download_fastqs
from protospacex import get_cds_chr_loc, get_cds_seq, get_ultramer_seq

from main import samplesheet
from main.forms import *
from main.models import *
from utils import conversions
from utils.validators import is_ensemble_transcript

logger = logging.getLogger(__name__)


class IndexView(ListView):
    model = Experiment
    template_name = 'index.html'

    def get_context_data(self, **kwargs):
        try:
            context = super().get_context_data(**kwargs)
            context['analyses'] = Analysis.objects.filter(owner=self.request.user)
            context['experiments'] = Experiment.objects.filter(owner=self.request.user)
            return context
        except ObjectDoesNotExist as e:
            raise Http404(e)


# TODO (gdingle): login automatically on first try?
# TODO (gdingle): email confirmation?
class CreateUserView(CreateView):
    template_name = 'base.html'
    form_class = CustomUserCreationForm
    success_url = '/main/'


class CreatePlusView(CreateView):
    """
    Simplifies adding foreign keys and other pre-determined data to a ModelForm
    before saving it.

    See https://github.com/django/django/blob/master/django/views/generic/edit.py.
    """

    def form_valid(self, form: ModelForm) -> HttpResponse:
        try:
            obj = form.save(commit=False)
            obj.owner = self.request.user

            obj = self.plus(obj)
            obj.full_clean()
        except (ValidationError, ValueError) as e:
            e2 = copy.copy(e)
            e2.args += (obj.owner.username, obj.owner.email)
            logging.exception(e2)
            # Force field specific errors to __all__ because they are likely
            # from fields excluded from the form.
            if isinstance(e, ValidationError) and hasattr(e, 'error_dict'):
                e.error_list = [e.message for el in e.error_dict.values() for e in el]
                del e.error_dict
            form.add_error('__all__', e)
            return self.form_invalid(form)
        # TODO (gdingle): test me
        except ObjectDoesNotExist as e:
            raise Http404(e)
        obj.save()
        self.object = obj
        return HttpResponseRedirect(self.get_success_url())

    def plus(self, obj: models.Model) -> models.Model:
        """Set more fields a model before saving it."""
        return obj


class BaseDeleteView(SingleObjectMixin, View):
    """
    Works like DeleteView, but without confirmation screens or a success_url.
    """
    pk_url_kwarg = 'id'

    def post(self, *args, **kwargs):
        self.object = self.get_object()
        if self.object.owner.id != self.request.user.id:
            return HttpResponseForbidden()
        self.object.delete()
        return HttpResponseRedirect('/main/')


class ExperimentView(CreatePlusView):
    template_name = 'experiment.html'
    form_class = ExperimentForm
    success_url = '/main/experiment/{id}/guide-design/'


class ExperimentDeleteView(BaseDeleteView):
    model = Experiment


class GuideDesignView(CreatePlusView):
    template_name = 'guide-design.html'
    success_url = '/main/guide-design/{id}/progress/'

    def get_form_class(self):
        try:
            experiment = Experiment.objects.get(
                owner=self.request.user, id=self.kwargs['id'])
        except ObjectDoesNotExist as e:
            raise Http404(e)

        # Filter irrelevant fields depending on is_hdr
        if experiment.is_hdr:
            return GuideDesignForm
        else:
            return GuideDesignForm2

    def _get_targets_chr_loc(self, targets, guide_design):
        genome = guide_design.genome

        if all(is_chr(t) for t in targets):
            return targets

        elif guide_design.is_hdr:
            if not all(is_ensemble_transcript(t) or is_gene(t) for t in targets):
                raise ValidationError(
                    'Targets must all be ENST transcripts or gene symbols for HDR')
            if genome != 'hg38':
                raise ValidationError(
                    'ENST transcripts are only currently implemented for the hg38 genome')
            if guide_design.hdr_tag == 'per_target':
                func = get_cds_chr_loc
                cds_indexes = guide_design.cds_index
                if not cds_indexes:
                    raise ValidationError('You must specify "N" or "C" for each target')
                cds_lengths = guide_design.cds_length
                with ThreadPoolExecutor(4) as pool:
                    return list(pool.map(func, targets, cds_indexes, cds_lengths))
            else:
                func = functools.partial(
                    get_cds_chr_loc,
                    cds_index=guide_design.cds_index,
                    length=guide_design.cds_length)
                with ThreadPoolExecutor(4) as pool:
                    return list(pool.map(func, targets))

        elif all(is_seq(t) for t in targets):
            func = functools.partial(
                conversions.seq_to_chr_loc,
                genome=genome)
            with ThreadPoolExecutor(4) as pool:
                return list(pool.map(func, targets))

        elif all(is_gene(t) or is_ensemble_transcript(t) for t in targets):
            # TODO (gdingle): this still needs some work to get best region of gene and not the whole thing
            func = functools.partial(
                conversions.gene_to_chr_loc,
                genome=genome)
            with ThreadPoolExecutor(4) as pool:
                return list(pool.map(func, targets))

        raise ValidationError('Targets must be all of one accepted type')

    def _get_target_seqs(self, targets, guide_design):
        if all(is_seq(t) for t in targets):
            return targets

        genome = guide_design.genome

        if guide_design.is_hdr:
            if genome != 'hg38':
                raise ValidationError(
                    'ENST transcripts are only currently implemented for the hg38 genome')
            if guide_design.hdr_tag == 'per_target':
                func = get_cds_seq
                cds_indexes = guide_design.cds_index
                cds_lengths = guide_design.cds_length
                with ThreadPoolExecutor(4) as pool:
                    return list(pool.map(func, targets, cds_indexes, cds_lengths))
            else:
                func = functools.partial(
                    get_cds_seq,
                    cds_index=guide_design.cds_index,
                    length=guide_design.cds_length)
                with ThreadPoolExecutor(4) as pool:
                    return list(pool.map(func, targets))

        elif all(is_ensemble_transcript(t) for t in targets):
            raise ValidationError(
                'ENST transcripts are only currently implemented for HDR')

        elif all(is_gene(t) for t in targets):
            raise ValidationError(
                'Targeting genes by name is only currently implemented for HDR')

        elif all(is_chr(t) for t in targets):
            func = functools.partial(
                conversions.chr_loc_to_seq,
                genome=genome)
            with ThreadPoolExecutor(4) as pool:
                return list(pool.map(func, targets))

        raise ValidationError('Targets must be all of one accepted type')

    def _get_target_genes(self, targets, guide_design):
        if all(is_gene(t) for t in targets):
            return targets

        elif all(is_ensemble_transcript(t) for t in targets):
            func = functools.partial(
                conversions.enst_to_gene_or_unknown,
                genome=guide_design.genome)

        elif all(is_seq(t) for t in targets):
            func1 = functools.partial(
                conversions.seq_to_chr_loc,
                genome=guide_design.genome)
            func2 = functools.partial(
                conversions.chr_loc_to_gene,
                genome=guide_design.genome)

            def func(target):  # type: ignore
                return func2(func1(target))

        elif all(is_chr(t) for t in targets):
            func = functools.partial(
                conversions.chr_loc_to_gene,
                genome=guide_design.genome)

        else:
            raise ValidationError('Targets must be all of one accepted type')

        # More than 4 starts causing strange 404 errors from togows.org
        # and it causes 'You have exceeded the limit' errors from UCSC :(
        with ThreadPoolExecutor(4) as pool:
            return list(pool.map(func, targets))

    def plus(self, obj):
        obj.experiment = Experiment.objects.get(
            owner=self.request.user, id=self.kwargs['id'])

        targets_cleaned, target_tags = obj.parse_targets_raw()
        obj.target_tags = target_tags

        logger.info('Getting chromosome locations...')
        obj.target_locs = self._get_targets_chr_loc(
            targets_cleaned,
            obj
        )
        logger.info('Getting target sequences...')
        obj.target_seqs = self._get_target_seqs(
            targets_cleaned,
            obj
        )
        logger.info('Getting target genes...')
        obj.target_genes = self._get_target_genes(
            targets_cleaned,
            obj
        )

        batch = webscraperequest.CrisporGuideBatchWebRequest(obj)

        largs = [[
            target_seq,
            obj.experiment.name,
            obj.genome,
            obj.pam,
            target,
            obj.pre_filter]
            for target, target_seq in zip(obj.target_locs, obj.target_seqs)]
        logger.info('Getting guides from Crispor...')
        batch.start(largs, [-2])

        return obj


class GuideDesignProgressView(View):

    template_name = 'guide-design-progress.html'
    success_url = '/main/guide-design/{id}/guide-selection/'

    def get(self, request, **kwargs):
        guide_design = GuideDesign.objects.get(
            owner=self.request.user, id=self.kwargs['id'])
        batch_status = webscraperequest.CrisporGuideBatchWebRequest(guide_design).get_batch_status()

        other_recent_usage = guide_design.other_recent_usage

        if not batch_status.is_successful:
            return render(request, self.template_name, locals())
        else:
            # For unknown reason, redirecting too early causes 500 error
            time.sleep(1)
            return HttpResponseRedirect(
                self.success_url.format(id=self.kwargs['id']))


class GuideSelectionView(CreatePlusView):
    template_name = 'guide-selection.html'
    form_class = GuideSelectionForm
    success_url = '/main/guide-selection/{id}/primer-design/'

    def _get_top_guides(self, guide_design, min_score=10, max_hdr_dist=40) -> dict:
        """
        Filters all guides returned by Crispor down to those that have a score
        greater than min_score, then takes top guides by special ranking
        if HDR, else by score.

        Doench min_score = 10 is the same value used by Benchling.

        max_hdr_dist is based on effective limits of HDR.
        """
        # TODO (gdingle): figure out why guide_seqs not always there!
        selected_guides = dict(
            (g['target'], g.get('guide_seqs'))
            for g in guide_design.guide_data
            if webscraperequest.NOT_FOUND not in g.get('guide_seqs'))
        not_founds = dict(
            (g['target'], g.get('guide_seqs'))
            for g in guide_design.guide_data
            if webscraperequest.NOT_FOUND in g.get('guide_seqs'))

        if not selected_guides:
            # Early exit for edge case
            return not_founds

        # Make temp obj for samplesheet
        guide_selection = GuideSelection(
            guide_design=guide_design,
            selected_guides=selected_guides
        )

        # TODO (gdingle): remove me when calibration done
        from utils import manuscore
        manuscore._specificity_weight_low = int(self.request.GET.get('sl', 45))
        manuscore._specificity_weight_high = int(self.request.GET.get('sh', 65))
        manuscore._dist_weight_variance = int(self.request.GET.get('dv', 55))
        manuscore._before_start_codon_penalty = float(self.request.GET.get('cp', 0.2))

        sheet = samplesheet.from_guide_selection(guide_selection)
        sheet = sheet.loc[sheet['guide_score'] >= min_score, :]
        if not len(sheet):
            # TODO (gdingle): handle zero guides case better
            raise ValueError('No good guides found for any targets')

        if guide_design.is_hdr:
            sheet = sheet.loc[sheet['hdr_dist'] <= max_hdr_dist, :]
            sheet.sort_values('hdr_score', inplace=True, ascending=False)
        else:
            sheet.sort_values('guide_score', inplace=True, ascending=False)

        grouped = sheet.groupby(['target_loc'])
        # Take top then regroup for iteration by group
        top = guide_design.guides_per_target
        grouped = grouped.head(top).groupby(['target_loc'])
        top_guides = dict((
            str(target_loc),
            dict(zip(group['_crispor_pam_id'],
                     group['guide_seq'] + ' ' + group['guide_pam'])))
            for target_loc, group in grouped)
        # Place not founds at top for easy reading
        return {**not_founds, **top_guides}

    def get_initial(self):
        guide_design = GuideDesign.objects.get(
            owner=self.request.user, id=self.kwargs['id'])
        return {
            'selected_guides': self._get_top_guides(guide_design),
        }

    def plus(self, obj):
        obj.guide_design = GuideDesign.objects.get(
            owner=self.request.user, id=self.kwargs['id'])
        return obj

    def get_context_data(self, **kwargs):
        try:
            guide_design = GuideDesign.objects.get(
                owner=self.request.user, id=self.kwargs['id'])
            # TODO (gdingle): try to order these in the same way as _get_top_guides
            kwargs['crispor_urls'] = guide_design.crispor_urls
            return super().get_context_data(**kwargs)
        except ObjectDoesNotExist as e:
            raise Http404(e)


class PrimerDesignView(CreatePlusView):
    template_name = 'primer-design.html'
    form_class = PrimerDesignForm
    success_url = '/main/primer-design/{id}/progress/'

    def get_form_kwargs(self):
        return {
            **super().get_form_kwargs(),
            **{'id': self.kwargs['id']},
        }

    def plus(self, obj):
        guide_selection = GuideSelection.objects.get(
            owner=self.request.user, id=self.kwargs['id'])
        obj.guide_selection = guide_selection

        sheet = samplesheet.from_guide_selection(guide_selection)
        batch = webscraperequest.CrisporPrimerBatchWebRequest(obj)

        largs = [[row['_crispor_batch_id'],
                  row['_crispor_pam_id'],
                  obj.max_amplicon_length,
                  obj.primer_temp,
                  guide_selection.guide_design.pam,
                  row['_guide_id'],
                  self._get_hdr_dist_for_crispor(row)]
                 for row in sheet.to_records()
                 if row['guide_seq']]
        batch.start(largs, [-2])

        # TODO (gdingle): run crispr-primer if HDR experiment
        # https://github.com/chanzuckerberg/crispr-primer
        return obj

    @staticmethod
    def _get_hdr_dist_for_crispor(row):
        """
        Presence of hdr_dist arg causes different primer design in Crispor.
        CrispyCrunch hdr_dist is relative to strand of gene.
        Crispor hdr_dist is relative to positive genome strand.
        """
        if 'hdr_dist' not in row.dtype.names:
            return None
        else:
            return int(row['hdr_dist']) * (
                1 if row['target_loc'].strand == '+' else -1)


class PrimerDesignProgressView(View):
    template_name = 'primer-design-progress.html'
    success_url = '/main/primer-design/{id}/primer-selection/'

    def get(self, request, **kwargs):
        primer_design = PrimerDesign.objects.get(
            owner=self.request.user, id=kwargs['id'])
        batch_status = webscraperequest.CrisporPrimerBatchWebRequest(
            primer_design).get_batch_status()
        other_recent_usage = primer_design.other_recent_usage

        if not batch_status.is_done:
            return render(request, self.template_name, locals())
        else:
            if not batch_status.is_successful:
                logger.warn('Advancing to PrimerSelectionView with errors')
                # TODO (gdingle): remove errors
            # For unknown reason, redirecting too early causes 500 error
            time.sleep(1)
            return HttpResponseRedirect(
                self.success_url.format(id=self.kwargs['id']))


class PrimerSelectionView(CreatePlusView):
    template_name = 'primer-selection.html'
    form_class = PrimerSelectionForm
    success_url = '/main/primer-selection/{id}/experiment-summary/'

    primer_product_min = 250
    primer_product_max = 310

    def get_initial(self):
        return {'selected_primers': self._selected_primers()}

    def get_context_data(self, **kwargs):
        try:
            primer_design = PrimerDesign.objects.get(
                owner=self.request.user, id=self.kwargs['id'])
            kwargs['crispor_urls'] = primer_design.crispor_urls
            kwargs['primerblast_urls'] = self._primerblast_urls(
                primer_design.guide_selection.guide_design)
            return super().get_context_data(**kwargs)
        except ObjectDoesNotExist as e:
            raise Http404(e)

    def plus(self, obj):
        obj.primer_design = PrimerDesign.objects.get(
            owner=self.request.user, id=self.kwargs['id'])

        # Needed for primerblast
        self._set_custom_primers(
            obj.primer_design.guide_selection.guide_design,
            obj,
        )

        return obj

    def _selected_primers(self) -> dict:
        primer_data = PrimerDesign.objects.get(
            owner=self.request.user, id=self.kwargs['id']).primer_data
        return dict(
            (p['target'], p['ontarget_primers'])
            for p in primer_data
            # PrimerSelection may happen in face of errors
            if 'target' in p)

    def _set_custom_primers(self, guide_design: GuideDesign, obj: PrimerSelection) -> None:
        if not guide_design.is_hdr:
            # TODO (gdingle): make it work for non-HDR
            return None

        ultramer_seqs = self._get_ultramer_seqs(guide_design)

        for guide_id, ultramer_seq in zip(self._selected_primers(), ultramer_seqs):
            primers = obj.selected_primers.get(guide_id)
            if primers is None:
                # TODO (gdingle): should we still allow removing whole primers elements from dict?
                logging.warn('No primers for guide {}'.format(guide_id))
                continue
            assert len(primers), 'No primers for {}'.format(guide_id)
            assert len(primers) <= 3, 'Too many primers for {}'.format(guide_id)
            if len(primers) == 2:
                obj.selected_primers[guide_id].append(
                    self._get_primer_product(ultramer_seq, *primers))

    def _get_primer_product(self, ultramer_seq: str, primer_seq_fwd: str, primer_seq_rev: str) -> str:
        ultramer_seq = ultramer_seq.upper()
        left = ultramer_seq.index(primer_seq_fwd)
        right = ultramer_seq.rindex(reverse_complement(primer_seq_rev)) + len(primer_seq_rev)
        assert right - left <= self.primer_product_max
        assert right - left >= self.primer_product_min

        primer_product = ultramer_seq[left: right]
        assert primer_product.startswith(primer_seq_fwd)
        assert primer_product.endswith(reverse_complement(primer_seq_rev))
        return primer_product

    def _get_ultramer_seqs(self, guide_design: GuideDesign) -> list:

        def _get_ultramer_seq(target_input: str, cds_index: int, chr_loc: ChrLoc) -> Optional[str]:
            try:
                # TODO (gdingle): rename get_ultramer_seq to get_extended_seq or something
                ultramer_seq = get_ultramer_seq(
                    target_input,
                    cds_index,
                    # Must be long enough to find primers. Jason
                    # said "we’d need to extract the -205 to +205 region around
                    # the insert site and put that into PCR template box".
                    self.primer_product_max + 100)[0]
                if chr_loc.strand == '-':
                    return reverse_complement(ultramer_seq)
                else:
                    return ultramer_seq
            except Exception as e:
                logger.warn(e)
                return None

        targets_cleaned, _ = guide_design.parse_targets_raw()

        if guide_design.hdr_tag == 'per_target':
            cds_indexes = guide_design.cds_index
        else:
            cds_indexes = [guide_design.cds_index] * len(targets_cleaned)

        # TODO (gdingle): another instance of IO... how to avoid?
        with ThreadPoolExecutor(4) as pool:
            ultramer_seqs = list(pool.map(
                lambda args: _get_ultramer_seq(*args),
                zip(targets_cleaned, cds_indexes, guide_design.target_locs),
            ))
        return ultramer_seqs

    def _primerblast_urls(
        self,
        guide_design: GuideDesign,
        base_url='https://www.ncbi.nlm.nih.gov/tools/primer-blast/index.cgi'
    ) -> Optional[dict]:

        if not guide_design.is_hdr:
            # TODO (gdingle): implement for non-HDR, non-ENST
            return None

        ultramer_seqs = self._get_ultramer_seqs(guide_design)

        not_founds = dict(
            (p[0], ultramer_seqs[i])
            for i, p in enumerate(self._selected_primers().items())
            if p[1] == [webscraperequest.NOT_FOUND])

        params = {
            'LINK_LOC': 'bookmark',
            'PRIMER5_START': 1,
            'PRIMER5_END': 100,
            'PRIMER3_START': 311,
            'PRIMER3_END': 410,
            'PRIMER_PRODUCT_MIN': self.primer_product_min,
            'PRIMER_PRODUCT_MAX': self.primer_product_max,
            'PRIMER_SPECIFICITY_DATABASE': 'refseq_representative_genomes',
            'ORGANISM': guide_design.ncbi_organism,
            'PRIMER_MIN_SIZE': 15,
            'PRIMER_OPT_SIZE': 20,
            'PRIMER_MAX_SIZE': 25,
        }
        return dict(
            (target,
             base_url + '?' + urllib.parse.urlencode({
                 **params, **{'INPUT_SEQUENCE': ultramer_seq}
             }))
            for target, ultramer_seq in not_founds.items()
            # Simply filter out errors for now
            if ultramer_seq
        )


class ExperimentSummaryView(View):
    template_name = 'experiment-summary.html'

    def get(self, request, *args, **kwargs):
        # Two paths to this View... see urls.py
        try:
            experiment_id = kwargs.get('experiment_id')
            if experiment_id:
                primer_selection = PrimerSelection.objects.filter(
                    owner=self.request.user,
                    primer_design__guide_selection__guide_design__experiment=experiment_id)[0]
            else:
                primer_selection = PrimerSelection.objects.get(
                    owner=self.request.user, id=kwargs['primer_selection_id'])
        except (PrimerSelection.DoesNotExist, IndexError):
            raise Http404('Experiment summary does not exist')

        ms = request.GET.get('ms')
        sheet = samplesheet.from_primer_selection(primer_selection,
                                                  # TODO (gdingle): temp remove me,
                                                  float(ms) if ms else None,
                                                  False)
        if request.GET.get('download') == 'xls':
            # A couple of hidden cols needed for analysis
            # TODO (gdingle): different name convention for analysis only fields?
            sheet['primer_adapt_name'] = sheet['_primer_adapt_name']
            if '_primer_product_wt' in sheet:
                sheet['primer_product_wt'] = sheet['_primer_product_wt']

        sheet = self._prepare_sheet(sheet)

        primer_design = primer_selection.primer_design
        guide_selection = primer_design.guide_selection
        guide_design = guide_selection.guide_design
        experiment = guide_design.experiment

        if request.GET.get('download') == 'xls':
            del sheet['Well Pos']  # already contained in index
            excel_file = samplesheet.to_excel(sheet)
            title = experiment.name + ' summary'
            return _excel_download_response(excel_file, title)

        # TODO (gdingle): download csv?

        # max length to show of table cell values
        # 26 is optimized for laptop screen and chr loc
        show = request.GET.get('show', 26)

        return render(request, self.template_name, locals())

    def _prepare_sheet(self, sheet):
        """Modify sheet for optimal rendering"""
        sheet = sheet.loc[:, [not c.startswith('_') for c in sheet.columns]]

        # Remove redundant info in case of genes
        if all(sheet['target_gene'] == sheet['target_input']):
            sheet['target_gene'] = None

        # Remove redundant info in case of chr
        if all((is_chr(t) for t in sheet['target_input'])) \
                and all(sheet['target_loc'] == sheet['target_input']):
            sheet['target_loc'] = None

        # Drop cols that are all na
        sheet = sheet.dropna(axis=1, how='all')

        # Fill remaining cells that are na
        sheet['target_input'] = list(sheet['target_input'])
        sheet = sheet.fillna('')

        # Prettify col names
        sheet.columns = [c.replace('_', ' ').title() for c in sheet.columns]

        return sheet


class AnalysisView(CreatePlusView):
    template_name = 'analysis.html'
    form_class = AnalysisForm
    success_url = '/main/analysis/{id}/progress/'

    def get_context_data(self, **kwargs):
        try:
            context = super().get_context_data(**kwargs)
            owner_exps = Experiment.objects.filter(owner=self.request.user)
            special_exp = Experiment.objects.filter(id=1)
            # TODO (gdingle): legacy... remove when id=1
            legacy_exp = Experiment.objects.filter(name='No experiment -- Custom analysis')
            context['form'].fields['experiment'].queryset = owner_exps | special_exp | legacy_exp
            return context
        except ObjectDoesNotExist as e:
            raise Http404(e)

    def plus(self, obj):
        # TODO (gdingle): use predetermined s3 location of fastq
        fastqs = download_fastqs(obj.s3_bucket, obj.s3_prefix, overwrite=False)
        if len(fastqs) >= 384:
            raise ValueError('Fastqs should be from max one plate')

        # Redirect to intermediate page if custom analysis
        if obj.is_custom:
            self.success_url = '/main/analysis/{id}/custom/'
            obj.fastq_data = fastqs
            return obj

        sheet = samplesheet.from_analysis(obj)

        # TODO (gdingle): create dir per download, as in seqbot
        obj.fastq_data = find_matching_pairs(
            fastqs,
            sheet.to_records(),
            parallelize=True,
            demultiplex=obj.demultiplex,
        )

        sheet = samplesheet.from_analysis(obj)

        webscraperequest.CrispressoBatchWebRequest.start_analysis(
            obj, sheet.to_records())
        return obj


class AnalysisDeleteView(BaseDeleteView):
    model = Analysis


class CustomAnalysisView(View):

    template_name = 'custom-analysis.html'
    success_url = '/main/analysis/{id}/progress/'
    form = CustomAnalysisForm

    def get(self, request, **kwargs):
        analysis = Analysis.objects.get(
            owner=self.request.user, id=kwargs['id'])
        return render(request, self.template_name, {
            **kwargs,
            'form': self.form(),
            'analysis': analysis,
        })

    def post(self, request, **kwargs):
        analysis = Analysis.objects.get(
            owner=self.request.user, id=kwargs['id'])
        form = self.form(request.POST, request.FILES)

        if not form.is_valid():
            return render(request, self.template_name, {
                **kwargs,
                'form': self.form(),
                'analysis': analysis,
            })

        try:
            file = form.cleaned_data['file']
            sheet = samplesheet.from_excel(file)
            fastq_data = analysis.fastq_data
            assert len(fastq_data), 'Fastqs must be present for a custom analysis'

            # fastq_data is initially a flat list. Process only on initial post
            # because find_matching_pairs is expensive.
            if isinstance(fastq_data[0], str):
                fastq_data = find_matching_pairs(
                    fastq_data,
                    sheet.to_records(),
                    parallelize=True,
                    demultiplex=analysis.demultiplex,
                )
        except (ValidationError, ValueError) as e:
            form.add_error('__all__', e)
            return render(request, self.template_name, {
                **kwargs,
                'form': form,
                'analysis': analysis,
            })

        file = form.cleaned_data['file']
        sheet = samplesheet.from_excel(file)

        fastq_data = analysis.fastq_data
        sheet['fastq_fwd'] = fastq_data[0]
        sheet['fastq_rev'] = fastq_data[0]

        # Save only after sheet is validated above
        analysis.fastq_data = fastq_data
        analysis.save()

        webscraperequest.CrispressoBatchWebRequest.start_analysis(
            analysis, sheet.to_records())

        return HttpResponseRedirect(
            self.success_url.format(id=self.kwargs['id']))


class AnalysisProgressView(View):
    template_name = 'analysis-progress.html'
    success_url = '/main/analysis/{id}/results/'

    def get(self, request, **kwargs):
        analysis = Analysis.objects.get(
            owner=self.request.user, id=kwargs['id'])
        batch_status = webscraperequest.CrispressoBatchWebRequest(analysis).get_batch_status()

        other_recent_usage = analysis.other_recent_usage

        if not batch_status.is_successful:
            return render(request, self.template_name, locals())
        else:
            return HttpResponseRedirect(
                self.success_url.format(id=self.kwargs['id']))


class ResultsView(View):
    template_name = 'crispresso-results.html'

    def get(self, request, *args, **kwargs):
        analysis = Analysis.objects.get(
            owner=self.request.user, id=self.kwargs['id'])
        try:
            if analysis.is_custom:
                sheet = samplesheet.from_custom_analysis(analysis)
                input_data = [
                    (d['input_data'],
                        [f.split('/')[-1] for f in d['input_files']])
                    for d in analysis.results_data]
            else:
                sheet = samplesheet.from_analysis(analysis)
            # TODO (gdingle): mode to show each stat col as % of total?
        except IndexError:
            raise Http404('Analysis results do not exist')
        return render(request, self.template_name, locals())


# TODO (gdingle): is DetailView needed here?
class OrderFormView(DetailView):
    """
    Produces a downloadable Excel order form for IDT. The model must have a
    plate layout.
    """

    model: models.Model = None
    seq_key: str

    def _create_excel_file(self, sheet: samplesheet.pandas.DataFrame, title: str):
        # TODO (gdingle): use pandas to_excel here instead?
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = title[0:31]  # Excel limits to 30 chars
        ws['A1'] = 'Well Position'
        ws['B1'] = 'Sequence Name'
        ws['C1'] = 'Sequence'

        # TODO (gdingle): how to avoid?
        # '_guide_id' may be dropped on earlier join
        sheet = sheet.reset_index()

        for j, well_pos in enumerate(sheet.index):
            for i, seq_key in enumerate(self.seq_keys):
                # TODO (gdingle): what to do if empty?
                index = str((j * len(self.seq_keys)) + i + 2)
                # TODO (gdingle): is this a good name for each sequence?
                row = sheet.loc[well_pos]
                ws['A' + index] = row['well_pos']
                ws['B' + index] = '{} {}'.format(
                    row['_guide_id'], ''.join(k for k in seq_key),
                )
                # TODO (gdingle): bad polymorphism, I know :/
                if isinstance(seq_key, tuple):
                    # TODO (gdingle): how to avoid values such as
                    # 'AGACGTGTGCTCTTCCGATCTnot found' ?
                    ws['C' + index] = ''.join(row[k] for k in seq_key)
                else:
                    ws['C' + index] = row[seq_key]

        return openpyxl.writer.excel.save_virtual_workbook(wb)

    def get(self, request, *args, **kwargs):
        instance = self.model.objects.get(
            owner=self.request.user, id=kwargs['id'])
        # TODO (gdingle): friendlier title?
        title = request.path.replace('/', ' ').replace('main ', '')
        excel_file = self._create_excel_file(instance.samplesheet, title)
        return _excel_download_response(excel_file, title)


class GuideOrderFormView(OrderFormView):

    model = GuideSelection
    seq_keys = ('guide_seq',)


class PrimerOrderFormView(OrderFormView):

    model = PrimerSelection
    seq_keys = (
        ('_primer_adapt_seq_fwd', 'primer_seq_fwd'),
        ('_primer_adapt_seq_rev', 'primer_seq_rev'),
    )


class UltramerOrderFormView(OrderFormView):
    """
    For HDR donor template DNA.
    """
    model = PrimerSelection
    seq_keys = ('_hdr_ultramer',)


class IlluminaSheetView(View):

    def _make_sample(self, guide_design: GuideDesign, row: Any) -> illumina.Sample:
        experiment = guide_design.experiment
        return illumina.Sample({
            # A short ID assigned to the specific study and/or project based
            #  on the conventions used at your institution/group. Accepted
            #  characters include numbers, letters, "-", and "_".
            'Study_ID': experiment.short_name,

            # Description of a project to which the current sequencing run
            # belongs. Special characters and commas are not allowed.
            # (Required)
            # e.g. Understanding differences in RNA expression
            # between pancreatic cells collected from healthy pancreas and
            # patients with pancreatic cancer.
            'Study_Description': experiment.description,

            # Description (longer than 10 characters) of the specific
            # biological sample. The description should be the same for all
            # the samples derived from this BioSample. Special characters
            # and commas are not allowed.
            # (Required)
            # e.g. Microbiome sample collected on Feb. 30 2017 from mouse 002.
            # e.g. Liver tissue obtained on Feb. 30 2017 from patient 012.
            # Leave blank, to be filled in by researcher
            'BioSample_ID': None,

            # Description (longer than 10 characters) of the specific
            # biological sample. The description should be the same for all
            # the samples derived from this BioSample. Special characters
            # and commas are not allowed.
            # (Required) e.g. Microbiome sample
            # collected on Feb. 30 2017 from mouse 002. e.g. Liver tissue
            # obtained on Feb. 30 2017 from patient 012.
            # Leave blank, to be filled in by researcher
            'BioSample_Description': None,

            # A short ID assigned to the specific library based on the
            # conventions used at your institution/group. Accepted
            # characters include numbers, letters, "-", and "_". This field
            # should be unique for each row. Sample_ID can be the same as
            # Sample_Name
            # (Required)
            'Sample_ID': f'{experiment.short_name}-{row.index}',

            # A distinct and descriptive name for each specific library.
            # Accepted characters are numbers, letters, "-", and "_". Name
            # must begin with a letter. This field should be unique for each
            # row. The sample name will go into the final fastq file names.
            # (Required) e.g. Mouse_Liver_SingleCell_plate02_A10_20171210
            'Sample_Name': f'{experiment.short_name}-{row.index}',

            # If people are combining samples to sequence on the same run,
            # this column is used to keep track of to whom each sample
            # belongs to. Please fill in following the format
            # FirstName_LastName. If left blank, we will use submitter's
            # information.
            # e.g. Stephen_Quake
            'Sample_Owner': experiment.owner.username,

            # Name of the first index. Accepted characters are numbers,
            # letters, "-", and "_"
            # (Required)
            # e.g. Nextera - N700
            # Leave blank, to be filled in by researcher
            'Index_ID': None,

            # Sequence of the first index
            # (Required)
            # e.g. ATAGCGCT
            # Leave blank, to be filled in by researcher
            'Index': None,

            # Name of the second index. Accepted characters are numbers,
            # letters, "-", and "_"
            # (Required) e.g. Nextera - S500
            # Leave blank, to be filled in by researcher
            'Index2_ID': None,

            # Sequence of the second index
            # (Required) e.g. ATAGCGCT
            # Leave blank, to be filled in by researcher
            'Index2': None,

            # This field is used to record the organism the DNA comes from.
            # Some examples are Human, Mouse, Mosquito, Yeast, Bacteria,
            # etc.
            # (Required)
            'Organism': guide_design.organism,

            # TODO (gdingle): will these below ever be useful?
            # This field is used record the host of the organism where the
            # DNA originated, if applicable. For example, if the Organism is
            # Fungus and the Fungus is isolated from a Human fecal sample,
            # then the Host is Human. If the organism is Human, then host is
            # left blank.
            'Host': None,

            # Gender of the organism (if applicable) e.g. M or F
            'Gender': None,

            # Tissue origin of the particular sample (if applicable) e.g. Liver
            'Tissue_Source': None,

            # FACS markers used delimited by a semicolon ";" with no space.
            # Accepted characters are numbers, letters, "-" and "_". e.g.
            # Epcam;CD45
            'FACS_Markers': None,
        })

    def get(self, request, **kwargs):
        # TODO (gdingle): do we actually need primer selection here? or is guide enough?
        primer_selection = PrimerSelection.objects.get(
            owner=self.request.user, id=kwargs['id'])
        sheet = samplesheet.from_primer_selection(primer_selection)
        guide_design = primer_selection.primer_design.guide_selection.guide_design
        experiment = guide_design.experiment

        illumina_sheet = illumina.SampleSheet()
        # TODO (gdingle): Add link to Biohub submission form
        illumina_sheet.Header['CrispyCrunch'] = 'Please fill in the following required columns: BioSample_ID, BioSample_Description, Index_ID, Index, Index2_ID, Index2.'
        for row in sheet.to_records():
            illumina_sheet.add_sample(self._make_sample(guide_design, row))

        csv = StringIO()
        illumina_sheet.write(csv)
        filename = f'Illumina sample sheet for experiment {experiment.name}'
        response = HttpResponse(csv.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="{}.csv"'.format(filename)

        return response


class ExampleCustomAnalysisSheetView(View):
    """
    Produces a example Excel sheet for customization and upload to CrispyCrunch.
    """

    def get(self, request, *args, **kwargs):
        title = 'CrispyCrunch custom analysis'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title  # Excel limits to 30 chars

        headers = [
            ('A', 'target_input', 'ENST00000617316'),
            ('B', 'guide_seq', 'GACAGCGCTGAACTTCAGTT'),
            ('C', 'primer_seq_fwd', 'GGAGGAATGAGCAGCA'),
            ('D', 'primer_seq_rev', 'AGCCATATGAAAGAGA'),
            ('E', 'primer_product', 'GGAGGAATGAGCAGCAGACATGGGAGACGGATGAGTCTTTTAATAGAAAAACACACGTGCAACAGTATCAACACACATCTCTCGCAATGCTCACAGCGCTGAACTTTAACCACTTCCTGGACCTTGAAACAAAACTTCCAATCCGCCACCCATCATATCGGTAAAGGCCTTTTGCCACTCCTTGAAGTTGAGCTCGGTGTTCTTCACCTTGGGGGGTGGCCTGTGAGAGGAAGATAGGTGATGAAGGAGGGTCCCCAGGATCATGGCACTCGGGGTGCAAGGGACAGAGATGTCTGTCTTGGTGTATTGCTGGGCCCCTGCTCACCTGTACACTCCCACGACCACGGCATGGTCTCTTTCATATGGCT'),
            # TODO (gdingle): rename to amplicon?
            # TODO (gdingle): add only if experiment is_hdr!!!
            ('F', 'primer_product_wt', 'GGAGGAATGAGCAGCAGACATGGGAGACGGATGAGTCTTTTAATAGAAAAACACACGTGCAACAGTATCAACACACATCTCTCGCAATCCTGACAGCGCTGAACTTCAGTTCTTCACCTTGGGGGGTGGCCTGTGAGAGGAAGATAGGTGATGAAGGAGGGTCCCCAGGATCATGGCACTCGGGGTGCAAGGGACAGAGATGTCTGTCTTGGTGTATTGCTGGGCCCCTGCTCACCTGTACACTCCCACGACCACGGCATGGTCTCTTTCATATGGCT'),
            ('G', 'primer_adapt_name', 'TruSeq3-PE.fa',),
        ]
        for cell, header, value in headers:
            ws[cell + '1'] = header
            ws[cell + '2'] = value

        excel_file = openpyxl.writer.excel.save_virtual_workbook(wb)
        return _excel_download_response(excel_file, title)


def _excel_download_response(excel_file: BytesIO, title: str) -> HttpResponse:
    response = HttpResponse(
        excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(title)
    return response
